import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

st.set_page_config(page_title="수시지불 대조 프로그램", page_icon="📊", layout="wide")

st.title("📊 수시지불 대조 및 이체파일 생성기 (Ver.8.3)")

# ⭐️ 불필요한 설명 텍스트 삭제 및 깔끔한 라디오 버튼만 남김
version_choice = st.radio(
    "법인 선택",
    ("DH", "YK"),
    horizontal=True,
    label_visibility="collapsed" 
)
is_yk = version_choice == "YK"

col1, col2 = st.columns(2)
with col1:
    acc_file = st.file_uploader("1️⃣ 수시지불리스트 (회계팀 파일, 필수)", type=['xls', 'xlsx'])
with col2:
    sap_file = st.file_uploader("2️⃣ SAP 이체파일 (선택사항, 수시지불.xls)", type=['xls', 'xlsx'])

def clean_name(text):
    if pd.isna(text) or str(text).strip().lower() in ['nan', 'none']: 
        return ""
    text = str(text).replace(' ', '').lower()
    text = re.sub(r'\(주\)|주식회사|\(재\)|\(사\)|\(유\)', '', text)
    text = text.replace('대학교', '대') 
    return text

if acc_file:
    if st.button(f"🚀 {version_choice} 대조 및 이체파일 생성", use_container_width=True):
        with st.spinner("데이터를 분석하고 이체파일을 생성 중입니다..."):
            try:
                # ==========================================
                # 1. 회계팀 데이터 전처리
                # ==========================================
                acc_df = pd.read_excel(acc_file, header=1, dtype=str)
                acc_df = acc_df.dropna(subset=['업체명'])
                acc_df['원본_순서'] = acc_df.index 
                
                acc_df['은행명'] = acc_df['은행명'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '').str.strip()
                if '은행코드' in acc_df.columns:
                    acc_df['은행코드'] = acc_df['은행코드'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '').str.strip()
                else:
                    acc_df['은행코드'] = ""
                    
                acc_df['원래계좌번호'] = acc_df['계좌번호'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '').str.strip()
                acc_df['비교용_계좌'] = acc_df['원래계좌번호'].str.replace(r'\D', '', regex=True)
                acc_df['회계팀금액'] = pd.to_numeric(acc_df['금액'], errors='coerce').fillna(0).astype(int)

                empty_mask = (acc_df['비교용_계좌'] == '') | (acc_df['은행명'] == '')
                acc_df.loc[empty_mask, '비교용_계좌'] = 'EMPTY_' + acc_df[empty_mask].index.astype(str)

                # ==========================================
                # 2. SAP 파일 존재 여부에 따른 분기 처리
                # ==========================================
                if sap_file:
                    sap_raw = pd.read_excel(sap_file, header=None, dtype=str)
                    header_row_index = 0
                    for i in range(min(10, len(sap_raw))):
                        if '계좌번호' in sap_raw.iloc[i].astype(str).tolist():
                            header_row_index = i
                            break
                            
                    sap_df = pd.read_excel(sap_file, header=header_row_index, dtype=str)
                    sap_df = sap_df.dropna(how='all') 
                    
                    sap_df['비교용_계좌'] = sap_df['계좌번호'].astype(str).str.replace(r'\D', '', regex=True)
                    sap_df['SAP금액'] = pd.to_numeric(sap_df['금액'], errors='coerce').fillna(0).astype(int)
                    sap_df['SAP예금주'] = sap_df['예금주'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '').str.strip()
                    if '입금계좌표시내용' not in sap_df.columns: sap_df['입금계좌표시내용'] = ''
                    
                    sap_df['Match_Key'] = sap_df['SAP예금주'].apply(clean_name)
                    sap_df['Match_Key'] = np.where(sap_df['Match_Key'] == '', 'AMT_' + sap_df['SAP금액'].astype(str), sap_df['Match_Key'])
                    sap_keys = set(sap_df['Match_Key'].dropna())

                    sap_grouped = sap_df.groupby('Match_Key').agg(
                        SAP_총금액=('SAP금액', 'sum'),
                        SAP_은행명=('은행명', 'first'),
                        SAP_계좌번호=('계좌번호', 'first'),
                        SAP_비교용계좌=('비교용_계좌', 'first'),
                        SAP_예금주=('SAP예금주', 'first'),
                        SAP_표시내용=('입금계좌표시내용', 'first')
                    ).reset_index()

                    def acc_match_key(row):
                        ven = clean_name(row['업체명'])
                        dep = clean_name(row['예금주'])
                        acc = row['비교용_계좌']
                        amt_key = 'AMT_' + str(row['회계팀금액'])
                        
                        if dep in sap_keys: return dep      
                        if ven in sap_keys: return ven      
                        if acc in sap_keys: return acc      
                        if amt_key in sap_keys: return amt_key 
                        if '가상계좌' in dep or not dep: return ven
                        return dep

                    acc_df['Match_Key'] = acc_df.apply(acc_match_key, axis=1)
                    acc_df['Match_Key'] = np.where(acc_df['Match_Key'] == '', acc_df['비교용_계좌'], acc_df['Match_Key'])
                    acc_totals = acc_df.groupby('Match_Key')['회계팀금액'].sum().reset_index(name='회계팀_총금액')
                    acc_df = pd.merge(acc_df, acc_totals, on='Match_Key', how='left')

                    merged = pd.merge(acc_df, sap_grouped, on='Match_Key', how='outer')

                else:
                    merged = acc_df.copy()
                    merged['SAP_은행명'] = ""
                    merged['SAP_계좌번호'] = ""
                    merged['SAP_총금액'] = np.nan
                    merged['SAP_예금주'] = ""
                    merged['SAP_표시내용'] = ""
                    merged['SAP_비교용계좌'] = ""
                    merged['회계팀_총금액'] = merged['회계팀금액']

                # ==========================================
                # 3. 데이터 정리 및 공통 포맷팅
                # ==========================================
                merged['최종_은행명'] = merged['은행명'].fillna(merged['SAP_은행명']).astype(str).replace(['nan', 'None', 'none', 'NaN'], '')
                merged['최종_은행코드'] = merged['은행코드'].fillna(merged['SAP_은행명']).astype(str).replace(['nan', 'None', 'none', 'NaN'], '')
                
                merged['최종_계좌번호'] = merged['원래계좌번호'].fillna(merged['SAP_계좌번호']).astype(str).replace(['nan', 'None', 'none', 'NaN'], '')
                merged['최종_금액'] = merged['회계팀금액'].fillna(merged['SAP_총금액']).astype(int)
                
                merged['최종_예금주'] = np.where(
                    merged['SAP_예금주'].isna() | (merged['SAP_예금주'] == ''),
                    merged['예금주'],
                    merged['SAP_예금주']
                )
                merged['최종_예금주'] = merged['최종_예금주'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '')
                
                merged['최종_표시내용'] = merged['SAP_표시내용'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '')
                merged['최종_CMS'] = "" 
                
                merged['YK_입금통장표시'] = "와이케이스틸"
                merged['출금통장표시'] = ""
                merged['메모'] = ""
                merged['휴대폰번호'] = ""

                merged['[대조]회계팀_금액'] = merged['회계팀금액'].fillna(0).astype(int)
                merged['[대조]회계팀_예금주'] = merged['예금주'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '')
                merged['[대조]회계팀_업체명'] = merged['업체명'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '')
                merged['[대조]적요'] = merged['적요'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '')
                merged['[대조]담당자'] = merged['담당자'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '')
                
                if '계좌타입' in merged.columns:
                    merged['[대조]계좌타입'] = merged['계좌타입'].astype(str).replace(['nan', 'None', 'none', 'NaN'], '').str.strip()
                else:
                    merged['[대조]계좌타입'] = ""

                # ==========================================
                # 4. 상태값(검증) 로직
                # ==========================================
                def check_status(row):
                    sap_tot = row['SAP_총금액']
                    acc_tot = row['회계팀_총금액']
                    
                    if not sap_file:
                        return '단독 변환 (SAP 없음)'
                        
                    match_key = str(row.get('Match_Key', ''))
                    
                    if pd.isna(sap_tot) or sap_tot == 0:
                        return 'SAP 누락'
                    if pd.isna(acc_tot) or acc_tot == 0:
                        return '회계팀 누락'
                    if sap_tot != acc_tot:
                        return '금액 불일치 (총액)'
                    
                    if match_key.startswith('AMT_'):
                        return '확인필요 (금액매칭)'
                        
                    sap_acc = str(row['SAP_비교용계좌'])
                    acc_acc = str(row['비교용_계좌'])
                    
                    if sap_acc != acc_acc:
                        return '정상 (회계팀 계좌적용)'
                    
                    return '정상'
                    
                merged['검증결과'] = merged.apply(check_status, axis=1)

                if '원본_순서' in merged.columns:
                    merged['원본_순서'] = merged['원본_순서'].fillna(999999)
                    merged = merged.sort_values(by='원본_순서', ascending=True).reset_index(drop=True)

                # ==========================================
                # 5. 버전에 따른 출력 열(Columns) 선택
                # ==========================================
                if is_yk:
                    final_report = merged[[
                        '최종_은행명', '최종_계좌번호', '최종_금액', '최종_예금주', 'YK_입금통장표시', '출금통장표시', '메모', '최종_CMS', '휴대폰번호',
                        '[대조]회계팀_금액', '[대조]회계팀_예금주', '[대조]회계팀_업체명', '[대조]적요', '[대조]담당자', '[대조]계좌타입', '검증결과'
                    ]]
                    final_report.columns = [
                        '입금은행', '입금계좌번호', '이체금액', '예상예금주', '입금통장표시', '출금통장표시', '메모', 'CMS코드', '받는분휴대폰번호',
                        '[대조]회계팀_금액', '[대조]회계팀_예금주', '[대조]회계팀_업체명', '[대조]적요', '[대조]담당자', '[대조]계좌타입', '[대조]검증상태'
                    ]
                else:
                    final_report = merged[[
                        '최종_은행명', '최종_계좌번호', '최종_금액', '최종_표시내용', '최종_예금주', '최종_CMS',
                        '[대조]회계팀_금액', '[대조]회계팀_예금주', '[대조]회계팀_업체명', '[대조]적요', '[대조]담당자', '[대조]계좌타입', '검증결과'
                    ]]
                    final_report.columns = [
                        '은행명', '계좌번호', '금액', '입금계좌표시내용', '예금주', 'CMS코드',
                        '[대조]회계팀_금액', '[대조]회계팀_예금주', '[대조]회계팀_업체명',
                        '[대조]적요', '[대조]담당자', '[대조]계좌타입', '[대조]검증상태'
                    ]

                wb = Workbook()
                ws = wb.active
                ws.title = "이체파일_업로드용"

                grey_fill = PatternFill(start_color='BDBDBD', end_color='BDBDBD', fill_type='solid') 
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                blue_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
                grey_light_fill = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')

                start_row = 1 if is_yk else 5
                data_start_row = 2 if is_yk else 6

                headers = list(final_report.columns)
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

                for row_num, row_data in enumerate(final_report.values, data_start_row):
                    type_idx = 14 if is_yk else 11
                    is_w_type = str(row_data[type_idx]).strip().lower() == 'w'
                    
                    for col_num, value in enumerate(row_data, 1):
                        amt_cols = [3, 10] if is_yk else [3, 7]
                        if col_num in amt_cols:
                            cell = ws.cell(row=row_num, column=col_num, value=int(value) if pd.notna(value) and value != '' else 0)
                            cell.number_format = '#,##0'
                        else:
                            cell = ws.cell(row=row_num, column=col_num, value=value)
                            
                        status_col = 16 if is_yk else 13
                        if col_num == status_col: 
                            if value == '정상':
                                cell.fill = green_fill
                            elif value == '정상 (회계팀 계좌적용)':
                                cell.fill = blue_fill 
                            elif '확인필요' in str(value) or value == '예금주 상이':
                                cell.fill = yellow_fill 
                            elif value == '단독 변환 (SAP 없음)':
                                cell.fill = grey_light_fill
                            else:
                                cell.fill = red_fill
                        elif is_w_type: 
                            cell.fill = grey_fill

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column].width = min((max_length + 2) * 1.2, 50) 

                final_output = io.BytesIO()
                wb.save(final_output)
                final_output.seek(0)
                
                corp_name = version_choice
                if not sap_file:
                    st.warning(f"⚠️ SAP 파일 없이 회계팀 파일만으로 {corp_name} 이체파일 형식을 생성했습니다.")
                else:
                    st.success(f"✅ {corp_name} 대조 및 생성 완료!")
                
                st.dataframe(final_report, use_container_width=True)
                
                st.download_button(
                    label=f"📥 {corp_name} 최종본 이체파일 다운로드",
                    data=final_output.getvalue(),
                    file_name=f"은행업로드_수시지불_최종본_ver8.3_{corp_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"오류가 발생했습니다: {e}")